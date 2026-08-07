"""
Flask Backend — PDF Document Reconciliation
Serves similarity data from SQLite (sqlite-vector) to the React SPA.
"""

import os
import sqlite3
import importlib.resources
from datetime import datetime
from io import BytesIO

import fitz  # PyMuPDF
from flask import Flask, jsonify, send_file, request
from flask_cors import CORS
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ── Paths & constants ──────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WEBAPP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "db", "pdf_reconciliation.db")
TABLE = "pdf_document_reconciliation"
COL = "embedding"
DIM = 1024

app = Flask(__name__)
CORS(app)


@app.route("/")
def index():
    """Serve the webapp HTML shell."""
    return send_file(os.path.join(WEBAPP_DIR, "webapp.html"), mimetype="text/html")


# ── Connection helper ──────────────────────────────────────────────────

def _conn():
    """Return a SQLite connection with sqlite-vector loaded and init'd."""
    c = sqlite3.connect(DB_PATH)
    ext = importlib.resources.files("sqlite_vector.binaries") / "vector"
    c.enable_load_extension(True)
    c.load_extension(str(ext))
    c.enable_load_extension(False)
    c.execute(f"SELECT vector_init('{TABLE}','{COL}','dimension={DIM},type=FLOAT32,distance=COSINE')")
    return c


# ── Routes ─────────────────────────────────────────────────────────────

@app.route("/api/documents")
def documents():
    c = _conn()
    rows = c.execute(f"SELECT id, document_name FROM {TABLE} ORDER BY id").fetchall()
    c.close()
    return jsonify({"documents": [{"id": r[0], "name": r[1]} for r in rows]})


@app.route("/api/similarity")
def similarity():
    c = _conn()

    # 1) Fetch all doc ids & names
    rows = c.execute(f"SELECT id, document_name, {COL} FROM {TABLE} ORDER BY id").fetchall()
    n = len(rows)
    ids   = [r[0] for r in rows]
    names = [r[1] for r in rows]

    # 2) Build N×N similarity matrix using vector_full_scan (cosine distance → similarity)
    matrix = [[0.0] * n for _ in range(n)]
    for i in range(n):
        emb = rows[i][2]
        dist_rows = c.execute(
            f"SELECT rowid, distance FROM vector_full_scan('{TABLE}','{COL}',?,?)",
            (emb, n)
        ).fetchall()
        dist_map = {r[0]: r[1] for r in dist_rows}
        for j in range(n):
            matrix[i][j] = round(1.0 - dist_map.get(ids[j], 1.0), 6)

    # 3) Top-10 matches per document (exclude self)
    top = {}
    for i in range(n):
        ranked = sorted(
            ((matrix[i][j], names[j]) for j in range(n) if j != i),
            reverse=True
        )
        top[names[i]] = [{"document": d, "score": round(s, 4)} for s, d in ranked[:10]]

    c.close()
    return jsonify({"doc_names": names, "matrix": matrix, "top_matches": top})


# ── PDF Report ─────────────────────────────────────────────────────────

# YlOrRd 9-class palette → RGB tuples for fpdf2
YLORRD_RGB = [
    (255, 255, 204), (255, 237, 160), (254, 217, 118), (254, 178, 76),
    (253, 141, 60),  (252, 78, 42),   (227, 26, 28),   (189, 0, 38),
    (128, 0, 38),
]


def _heat_rgb(score):
    """Map [0.5, 1.0] → YlOrRd index."""
    idx = min(8, max(0, int((score - 0.5) / 0.5 * 9)))
    return YLORRD_RGB[idx]


def _strip_ext(name):
    """Return filename without .pdf extension."""
    return name.replace(".pdf", "") if name.lower().endswith(".pdf") else name


def _get_char_count(doc_name: str) -> int:
    """Return character count for a document by reading its OCR text or PDF.

    doc_name is stored as a linux-style path relative to BASE_DIR,
    e.g. ``input/20260807_1430/subfolder/doc.pdf``.
    Prefers the mirrored OCR .txt under ``output/…``, falls back to
    extracting text from the source PDF via PyMuPDF.
    """
    rel = doc_name.replace("\\", "/")

    # ── Try mirrored OCR text: "input/X/Y/doc.pdf" → "output/X/Y/doc.txt"
    if rel.lower().endswith(".pdf"):
        core = rel[:-4]
    else:
        core = rel
    # Strip "input/" prefix if present, then re-prefix with "output/"
    if core.startswith("input/"):
        txt_rel = "output/" + core[6:] + ".txt"
    else:
        txt_rel = "output/" + core + ".txt"
    txt_full = os.path.join(BASE_DIR, txt_rel)
    if os.path.isfile(txt_full):
        try:
            with open(txt_full, encoding="utf-8") as f:
                return len(f.read())
        except Exception:
            pass

    # ── Fallback: extract text from PDF via PyMuPDF
    pdf_full = os.path.join(BASE_DIR, rel)
    if os.path.isfile(pdf_full):
        try:
            doc = fitz.open(pdf_full)
            text = "".join(page.get_text() for page in doc)
            doc.close()
            return len(text)
        except Exception:
            pass

    return 0


def _build_report_pdf():
    """Build the full report PDF and return bytes."""
    # ── Fetch data (same logic as /api/similarity) ─────────────────
    c = _conn()
    rows = c.execute(f"SELECT id, document_name, {COL} FROM {TABLE} ORDER BY id").fetchall()
    n = len(rows)
    ids = [r[0] for r in rows]
    names = [r[1] for r in rows]

    matrix = [[0.0] * n for _ in range(n)]
    for i in range(n):
        emb = rows[i][2]
        dist_rows = c.execute(
            f"SELECT rowid, distance FROM vector_full_scan('{TABLE}','{COL}',?,?)",
            (emb, n)
        ).fetchall()
        dist_map = {r[0]: r[1] for r in dist_rows}
        for j in range(n):
            matrix[i][j] = round(1.0 - dist_map.get(ids[j], 1.0), 6)

    top = {}
    for i in range(n):
        ranked = sorted(((matrix[i][j], names[j]) for j in range(n) if j != i), reverse=True)
        top[names[i]] = [(d, s) for s, d in ranked[:10]]
    c.close()

    # ── Pre-compute char counts for all documents ─────────────────
    char_map = {name: _get_char_count(name) for name in names}

    # ── Build PDF ─────────────────────────────────────────────────
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "PDF Document Reconciliation Report", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"{n} documents  ·  BAAI/bge-m3  ·  cosine similarity", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # ── Section 1: Document Matches ───────────────────────────────
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, "Document Matches (top 10 per document)", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Column header hint
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, "FILE_IN_QUESTION  |  CLOSEST_MATCH  |  SIM%  |  #CHARS_Q  |  #CHARS_M",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(2)

    for idx, name in enumerate(names, 1):
        display = _strip_ext(name)
        src_chars = char_map.get(name, 0)
        # Header per document
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 7, f"{idx}. {display}  ({src_chars:,} chars)", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        matches = top[name]
        if not matches:
            pdf.cell(0, 5, "  (no other documents)", new_x="LMARGIN", new_y="NEXT")
        for rank, (m_name, score) in enumerate(matches, 1):
            m_display = _strip_ext(m_name)
            m_chars = char_map.get(m_name, 0)
            pdf.cell(0, 5,
                     f"  {rank:>2}. {m_display}  [{score*100:5.1f}%]  "
                     f"[{src_chars:,} | {m_chars:,} chars]",
                     new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    # ── Section 2: Heatmap ────────────────────────────────────────
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, "Similarity Heatmap (YlOrRd)", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Layout: label column + N data columns
    label_w = 55   # width for row labels
    cell_w = 13    # width per heat cell
    cell_h = 6     # height per cell
    fs = 6         # font size for scores
    label_fs = 6   # font size for labels

    # Effective page width
    page_w = pdf.w - 2 * pdf.l_margin
    cols_fit = min(n, max(3, int((page_w - label_w) / cell_w)))  # clamp
    cell_w = (page_w - label_w) / cols_fit  # recalc exact

    # Column headers (rotated abbreviated labels)
    pdf.set_font("Helvetica", "", 5)
    y0 = pdf.get_y()
    x0 = pdf.get_x()
    for c in range(cols_fit):
        lbl = _strip_ext(names[c])
        # truncate to ~3 chars for header
        short = lbl[:3] if len(lbl) > 3 else lbl
        x = x0 + label_w + c * cell_w + cell_w / 2
        pdf.set_xy(x - 2, y0)
        pdf.cell(4, cell_h, short, align="C")

    pdf.set_xy(x0, y0 + cell_h)
    pdf.set_font("Helvetica", "", fs)

    for r in range(n):
        y = pdf.get_y()
        # Row label
        pdf.set_font("Helvetica", "", label_fs)
        pdf.set_xy(x0, y)
        row_lbl = _strip_ext(names[r])
        if len(row_lbl) > 24:
            row_lbl = row_lbl[:22] + ".."
        pdf.cell(label_w, cell_h, row_lbl, align="R")
        pdf.set_font("Helvetica", "", fs)

        for c in range(cols_fit):
            score = matrix[r][c]
            r_val, g_val, b_val = _heat_rgb(score)
            pdf.set_fill_color(r_val, g_val, b_val)
            # text color
            tc = (255, 255, 255) if score > 0.85 else (0, 0, 0)
            pdf.set_text_color(*tc)
            x = x0 + label_w + c * cell_w
            pdf.set_xy(x, y)
            pdf.cell(cell_w, cell_h, f"{score*100:.0f}", fill=True, align="C")
            pdf.set_text_color(0, 0, 0)
        pdf.ln(cell_h)

    return pdf.output()


@app.route("/api/report/pdf")
def report_pdf():
    pdf_bytes = _build_report_pdf()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"document_reconciliation_report_{timestamp}.pdf"
    return send_file(
        BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


# ── Excel Report ───────────────────────────────────────────────────────

@app.route("/api/report/excel")
def report_excel():
    """Generate an Excel workbook with columns:
    FILE_IN_QUESTION | CLOSEST_MATCHING_FILE | SIMILARITY_SCORE
    (10 rows per document for its top-10 matches)."""
    # Fetch data (same as /api/similarity)
    c = _conn()
    rows = c.execute(f"SELECT id, document_name, {COL} FROM {TABLE} ORDER BY id").fetchall()
    n = len(rows)
    ids = [r[0] for r in rows]
    names = [r[1] for r in rows]

    matrix = [[0.0] * n for _ in range(n)]
    for i in range(n):
        emb = rows[i][2]
        dist_rows = c.execute(
            f"SELECT rowid, distance FROM vector_full_scan('{TABLE}','{COL}',?,?)",
            (emb, n)
        ).fetchall()
        dist_map = {r[0]: r[1] for r in dist_rows}
        for j in range(n):
            matrix[i][j] = round(1.0 - dist_map.get(ids[j], 1.0), 6)

    top = {}
    for i in range(n):
        ranked = sorted(((matrix[i][j], names[j]) for j in range(n) if j != i), reverse=True)
        top[names[i]] = [(d, s) for s, d in ranked[:10]]
    c.close()

    # ── Pre-compute char counts for all documents ─────────────────
    char_map = {name: _get_char_count(name) for name in names}

    # ── Build workbook ────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Similarity Report"

    # Header style
    header_fill = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    headers = [
        "FILE_IN_QUESTION",
        "CLOSEST_MATCHING_FILE",
        "SIMILARITY_SCORE",
        "#_CHARS_FILE_IN_QUESTION",
        "#_CHARS_CLOSEST_MATCHING_FILE",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows — 10 per document
    for doc_name in names:
        display = _strip_ext(doc_name)
        src_chars = char_map.get(doc_name, 0)
        matches = top[doc_name]
        for m_name, score in matches:
            m_display = _strip_ext(m_name)
            m_chars = char_map.get(m_name, 0)
            ws.append([display, m_display, round(score, 4), src_chars, m_chars])

    # Format columns
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 28

    # Percentage format for score column
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "0.00%"

    # Number format (comma-separated) for char-count columns
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"document_reconciliation_report_{timestamp}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/health")
def health():
    try:
        c = _conn()
        ver = c.execute("SELECT vector_version()").fetchone()[0]
        c.close()
        return jsonify({"status": "ok", "vector_version": ver})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# ── Main ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"DB: {DB_PATH}  (exists={os.path.exists(DB_PATH)})")
    app.run(host="0.0.0.0", port=5000, debug=True)
